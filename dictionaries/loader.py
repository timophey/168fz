"""
Загрузчик словарей из файлов
"""

import json
import csv
from pathlib import Path
from typing import Set, Dict, List, Optional
from datetime import datetime

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class DictionaryLoader:
    """Загрузчик словарей различных форматов"""

    @staticmethod
    def load_from_json(filepath: Path) -> Dict:
        """Загрузка словаря из JSON файла"""
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
            result = {}
            
            # Если это список слов
            if isinstance(data, list):
                result['words'] = set(word.lower() for word in data if isinstance(word, str))
                result['mappings'] = {}
            # Если это словарь с полем 'words'
            elif isinstance(data, dict):
                if 'words' in data and isinstance(data['words'], list):
                    result['words'] = set(word.lower() for word in data['words'] if isinstance(word, str))
                # Если это просто словарь слов {слово: значение}
                else:
                    result['words'] = set(k.lower() for k in data.keys() if isinstance(k, str))
                
                # Извлекаем mappings если есть
                if 'mappings' in data and isinstance(data['mappings'], dict):
                    result['mappings'] = {k.lower(): v for k, v in data['mappings'].items()}
                else:
                    result['mappings'] = {}
            else:
                result['words'] = set()
                result['mappings'] = {}
            
            return result

    @staticmethod
    def load_from_csv(filepath: Path, column: str = 'word') -> Set[str]:
        """Загрузка словаря из CSV файла"""
        words = set()
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if column in row:
                    words.add(row[column].strip().lower())
        return words

    @staticmethod
    def load_from_txt(filepath: Path) -> Set[str]:
        """Загрузка словаря из текстового файла (по одному слову на строку)"""
        words = set()
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                word = line.strip().lower()
                if word:
                    words.add(word)
        return words

    @staticmethod
    def load_from_xlsx(filepath: Path) -> Dict:
        """
        Загрузка словаря из XLSX файла
        
        Ожидаемая структура:
        - word (обязательно)
        - russian_analog (опционально) - для создания mappings
        - category (опционально) - категория словаря
        - description (опционально) - описание словаря
        
        Файл может содержать несколько листов - все они будут прочитаны и объединены.
        
        Возвращает словарь с полями: words, mappings, category, description
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("Библиотека openpyxl не установлена. Установите: pip install openpyxl")
        
        wb = load_workbook(filename=str(filepath), read_only=True, data_only=True)
        
        # Берем заголовки из первого листа (предполагаем, что все листы имеют одинаковую структуру)
        if not wb.sheetnames:
            raise ValueError("XLSX файл не содержит листов")
        
        first_ws = wb[wb.sheetnames[0]]
        
        # Читаем заголовки из первой строки
        headers = []
        for cell in first_ws[1]:
            headers.append(str(cell.value).strip().lower() if cell.value else '')
        
        # Проверяем наличие обязательной колонки 'word'
        if 'word' not in headers:
            raise ValueError("В XLSX файле отсутствует обязательная колонка 'word'")
        
        word_idx = headers.index('word')
        has_analog = 'russian_analog' in headers
        analog_idx = headers.index('russian_analog') if has_analog else -1
        has_category = 'category' in headers
        category_idx = headers.index('category') if has_category else -1
        has_description = 'description' in headers
        desc_idx = headers.index('description') if has_description else -1
        
        words = set()
        mappings = {}
        category = None
        descriptions = []
        
        # Читаем данные со всех листов
        for ws in wb.worksheets:
            # Пропускаем пустые листы
            if ws.max_row < 2:
                continue
            
            # Читаем данные начи со второй строки
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                
                word = str(row[word_idx]).strip() if row[word_idx] else ''
                if not word:
                    continue  # пропускаем пустые строки
                
                word_lower = word.lower()
                words.add(word_lower)
                
                # Если есть аналог, добавляем в mappings
                if has_analog and analog_idx < len(row) and row[analog_idx]:
                    analog = str(row[analog_idx]).strip()
                    if analog:
                        mappings[word_lower] = analog
                
                # Собираем категории (берем первую непустую)
                if has_category and category_idx < len(row) and row[category_idx] and not category:
                    category = str(row[category_idx]).strip()
                
                # Собираем описания (можно несколько)
                if has_description and desc_idx < len(row) and row[desc_idx]:
                    desc = str(row[desc_idx]).strip()
                    if desc:
                        descriptions.append(desc)
        
        wb.close()
        
        result = {
            'words': words,
            'mappings': mappings
        }
        
        if category:
            result['category'] = category
        if descriptions:
            # Берем первое описание или объединяем
            result['description'] = descriptions[0]
        
        return result

    @staticmethod
    def save_to_xlsx(dictionary_data: Dict, filepath: Path) -> None:
        """
        Экспорт словаря в XLSX файл
        
        Args:
            dictionary_data: словарь с полями words, mappings, name, version, category, description
            filepath: путь для сохранения
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("Библиотека openpyxl не установлена. Установите: pip install openpyxl")
        
        from openpyxl import Workbook
        
        # Максимальное количество строк в одном листе Excel (1-indexed)
        # Excel ограничивает 1,048,576 строк на лист
        MAX_ROWS_PER_SHEET = 1048576
        # Вычитаем 1 для заголовка
        MAX_DATA_ROWS = MAX_ROWS_PER_SHEET - 1
        
        wb = Workbook()
        
        # Данные
        words = sorted(dictionary_data.get('words', []))
        mappings = dictionary_data.get('mappings', {})
        category = dictionary_data.get('category', '')
        description = dictionary_data.get('description', '')
        dict_name = dictionary_data.get('name', 'Словарь')
        base_title = dict_name if dict_name else 'Словарь'
        
        # Если слов немного - создаем один лист
        if len(words) <= MAX_DATA_ROWS:
            sheets = [(base_title, words)]
        else:
            # Нужно несколько листов. Разбиваем слова по алфавитным границам,
            # стараясь не разрывать группы слов с одинаковой первой буквой.
            
            # Вспомогательная функция для получения первой буквы слова
            def get_first_letter(word: str) -> str:
                """Возвращает первую букву слова в нижнем регистре, или '' если нет букв"""
                for char in word:
                    if char.isalpha():
                        return char.lower()
                return ''
            
            # Группируем слова по первой букве
            letter_groups = []
            current_letter = None
            current_group = []
            
            for word in words:
                first_letter = get_first_letter(word)
                if first_letter != current_letter:
                    if current_group:
                        letter_groups.append((current_letter, current_group))
                    current_group = [word]
                    current_letter = first_letter
                else:
                    current_group.append(word)
            
            if current_group:
                letter_groups.append((current_letter, current_group))
            
            # Теперь распределяем группы по листам
            sheets = []
            current_sheet_words = []
            current_sheet_count = 0
            first_letter_of_sheet = None
            
            for letter, group in letter_groups:
                group_size = len(group)
                
                # Если текущий лист пуст, начинаем новый
                if current_sheet_count == 0:
                    first_letter_of_sheet = letter
                
                # Проверяем, помещается ли вся группа в текущий лист
                if current_sheet_count + group_size <= MAX_DATA_ROWS:
                    # Добавляем всю группу в текущий лист
                    current_sheet_words.extend(group)
                    current_sheet_count += group_size
                else:
                    # Группа не помещается целиком
                    if current_sheet_count > 0:
                        # Текущий лист не пуст, завершаем его
                        last_word = current_sheet_words[-1]
                        last_letter = get_first_letter(last_word)
                        if first_letter_of_sheet and last_letter:
                            sheet_title = f"{base_title} {first_letter_of_sheet.upper()}-{last_letter.upper()}"
                        else:
                            sheet_title = f"{base_title} {len(sheets) + 1}"
                        sheets.append((sheet_title, current_sheet_words.copy()))
                    
                    # Начинаем новый лист с этой группы
                    # Если группа сама по себе больше максимума, придется разбить её на части
                    if group_size > MAX_DATA_ROWS:
                        # Разбиваем группу на части, которые помещаются
                        for i in range(0, group_size, MAX_DATA_ROWS):
                            chunk = group[i:i + MAX_DATA_ROWS]
                            chunk_first = get_first_letter(chunk[0]) if chunk else letter
                            chunk_last = get_first_letter(chunk[-1]) if chunk else letter
                            if i == 0:
                                # Первая часть этой группы
                                sheet_title = f"{base_title} {chunk_first.upper()}-{chunk_last.upper()}"
                            else:
                                # Последующие части
                                sheet_title = f"{base_title} {chunk_first.upper()}-{chunk_last.upper()} {len(sheets) + 1}"
                            sheets.append((sheet_title, chunk.copy()))
                        # После разбиения большой группы, текущий лист пуст
                        current_sheet_words = []
                        current_sheet_count = 0
                        first_letter_of_sheet = None
                    else:
                        # Группа помещается в новый лист целиком
                        current_sheet_words = group.copy()
                        current_sheet_count = group_size
                        first_letter_of_sheet = letter
            
            # Добавляем последний лист
            if current_sheet_words:
                last_word = current_sheet_words[-1]
                last_letter = get_first_letter(last_word)
                if first_letter_of_sheet and last_letter:
                    sheet_title = f"{base_title} {first_letter_of_sheet.upper()}-{last_letter.upper()}"
                else:
                    sheet_title = f"{base_title} {len(sheets) + 1}"
                sheets.append((sheet_title, current_sheet_words.copy()))
        
        # Заголовки
        headers = ['word', 'russian_analog', 'category', 'description']
        
        # Создаем листы и заполняем их
        for sheet_idx, (sheet_title, sheet_words) in enumerate(sheets):
            if sheet_idx == 0:
                ws = wb.active
                ws.title = sheet_title[:31]  # Excel ограничивает имя листа 31 символом
            else:
                ws = wb.create_sheet(title=sheet_title[:31])
            
            # Заголовки
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Данные
            for row_idx, word in enumerate(sheet_words, 2):
                ws.cell(row=row_idx, column=1, value=word)
                
                # Аналог если есть
                if word in mappings:
                    ws.cell(row=row_idx, column=2, value=mappings[word])
                
                # Категория (такая для всех строк)
                if category:
                    ws.cell(row=row_idx, column=3, value=category)
                
                # Описание (такое для всех строк)
                if description:
                    ws.cell(row=row_idx, column=4, value=description)
            
            # Автоматическая ширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        wb.close()

    @staticmethod
    def load_dictionary(filepath: Path) -> Dict:
        """
        Загрузка словаря с метаданными
        Возвращает словарь с полями: words, name, version, source, loaded_at, mappings (если есть)
        """
        if not filepath.exists():
            raise FileNotFoundError(f"Файл словаря не найден: {filepath}")

        ext = filepath.suffix.lower()

        loaders = {
            '.json': DictionaryLoader.load_from_json,
            '.csv': DictionaryLoader.load_from_csv,
            '.txt': DictionaryLoader.load_from_txt,
        }

        if ext not in loaders:
            raise ValueError(f"Неподдерживаемый формат файла: {ext}")

        result = loaders[ext](filepath)
        
        # Для JSON, CSV, TXT - result это set слов
        if isinstance(result, set):
            words = result
            mappings = {}
        else:
            # Если загрузчик вернул словарь (например, из XLSX)
            words = result.get('words', set())
            mappings = result.get('mappings', {})

        # Пытаемся извлечь метаданные из имени файла или соседних файлов
        name = filepath.stem
        version = "1.0"
        source = "local"
        category = None

        # Проверяем наличие файла с метаданными
        meta_file = filepath.with_suffix('.meta.json')
        if meta_file.exists():
            with open(meta_file, 'r', encoding='utf-8') as f:
                meta = json.load(f)
                name = meta.get('name', name)
                version = meta.get('version', version)
                source = meta.get('source', source)
                category = meta.get('category')

        # Если это JSON файл, пытаемся извлечь category, description и mappings из его содержимого
        if ext == '.json':
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, dict):
                        if 'category' in data and not category:
                            category = data['category']
                        # description уже может быть в result от load_from_json
                        if 'description' in data and 'description' not in result:
                            result['description'] = data['description']
            except:
                pass

        result = {
            'words': words,
            'name': name,
            'version': version,
            'source': source,
            'loaded_at': datetime.now().isoformat(),
            'filepath': str(filepath),
            'mappings': mappings
        }

        if category:
            result['category'] = category

        return result
